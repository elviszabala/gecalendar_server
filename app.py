from os import rename
from traceback import print_tb

import openpyxl

path = 'Horario.xlsm'

wb_obj = openpyxl.load_workbook(path, data_only=True)

d1 = [3,3]
d2 = [22,3]
d3 = [41,3]
d4 = [60,3]
d5 = [79,3]
d6 = [98,3]
d7= [117,3]
name = "Gevenis Narvaez"
week = [[3,3],[22,3],[41,3],[60,3],[79,3],[98,3],[117,3]]



sheet_obj = wb_obj.active

#cell_obj = sheet_obj.cell(row= 100 ,column=3)

max_colum = sheet_obj.max_column
print("Trabajador: ", name)
print("Semana: ", sheet_obj)

for k in range(0, len(week)):
    print("Dia: ", sheet_obj.cell(row=week[k][0], column=week[k][1]).value)
    for i in range(5, max_colum):

        cell_obj = sheet_obj.cell(row=week[k][0]+2, column=i)
        cell_hour = sheet_obj.cell(row=week[k][0], column=i)
        # print(cell_hour.value, end=" ")
        if(cell_obj.value is not None):
            print(cell_hour.value, cell_obj.value, end=" \n")



    print("Fin de dia --------------------------")




#print("CAntidad:", len(week))
