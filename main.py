from time import strptime

import datetime
from flask import Flask, request, jsonify
from flask_restful import Resource, Api
from flasgger import Swagger
from flask_swagger_ui import get_swaggerui_blueprint
import yaml
import os
import openpyxl
from flask_cors import CORS
from flask_socketio import SocketIO, send




path = 'Horario.xlsm'

wb_obj = openpyxl.load_workbook(path, data_only=True)

d1 = [3,3]
d2 = [22,3]
d3 = [41,3]
d4 = [60,3]
d5 = [79,3]
d6 = [98,3]
d7= [117,3]
name = "Gevenis Narvaez"
week = [[3,3],[22,3],[41,3],[60,3],[79,3],[98,3],[117,3]]

app = Flask(__name__)
api = Api(app)
CORS(app)
app.config['SECRET_KEY'] = 'mysecret'
socketio = SocketIO(app, cors_allowed_origins="*")

# Cargar la configuración de Swagger desde el archivo swagger.yaml
with open('static/swagger.yaml', 'r') as f:
    swagger_template = yaml.safe_load(f)

swagger = Swagger(app, template=swagger_template)



# Swagger config
SWAGGER_URL = '/swagger'
#API_URL = 'swagger.yaml'  # Aquí va el archivo JSON de la documentación de tu API
API_URL = '/static/swagger.yaml'  # Aquí va el archivo JSON de la documentación de tu API

swaggerui_blueprint = get_swaggerui_blueprint(
    SWAGGER_URL,  # Ruta donde Swagger estará disponible
    API_URL,      # Ruta del archivo swagger.json
    config={      # Configuración adicional para Swagger
        'app_name': "Flask API"
    }
)

# Registrar el blueprint de Swagger
app.register_blueprint(swaggerui_blueprint, url_prefix=SWAGGER_URL)

# Simulación de una base de datos en memoria
items = []
horario = []

# Rutas CRUD
class ItemList(Resource):
    def get(self):
        """Obtener todos los items
        ---
        responses:
          200:
            description: Lista de todos los items
        """
        return jsonify(items)

    def post(self):
        """Crear un nuevo item
        ---
        parameters:
          - in: body
            name: body
            description: Datos del nuevo item
            schema:
              type: object
              properties:
                name:
                  type: string
                  example: "Nuevo item"
        responses:
          201:
            description: Item creado
        """
        #print("Solicitud POST recibida")  # Agrega este print
        new_item = {
            'id': len(items) + 1,
            'name': request.json['name']
        }
        items.append(new_item)
        return jsonify(new_item), 201


class Item(Resource):
    def get(self, item_id):
        """Obtener un item por su ID
        ---
        parameters:
          - in: path
            name: item_id
            required: true
            schema:
              type: integer
        responses:
          200:
            description: Item encontrado
          404:
            description: Item no encontrado
        """
        item = next((item for item in items if item['id'] == item_id), None)
        if item:
            return jsonify(item)
        return jsonify({'message': 'Item no encontrado'}), 404

    def put(self, item_id):
        """Actualizar un item por su ID
        ---
        parameters:
          - in: path
            name: item_id
            required: true
            schema:
              type: integer
          - in: body
            name: body
            description: Datos para actualizar el item
            schema:
              type: object
              properties:
                name:
                  type: string
        responses:
          200:
            description: Item actualizado
          404:
            description: Item no encontrado
        """
        item = next((item for item in items if item['id'] == item_id), None)
        if item:
            item['name'] = request.json['name']
            return jsonify(item)
        return jsonify({'message': 'Item no encontrado'}), 404

    def delete(self, item_id):
        """Eliminar un item por su ID
        ---
        parameters:
          - in: path
            name: item_id
            required: true
            schema:
              type: integer
        responses:
          200:
            description: Item eliminado
          404:
            description: Item no encontrado
        """
        global items
        items = [item for item in items if item['id'] != item_id]
        return jsonify({'message': 'Item eliminado'})

def horas():
    sheet_obj = wb_obj['S43']
    #print(sheet_obj)
    # cell_obj = sheet_obj.cell(row= 100 ,column=3)

    max_colum = sheet_obj.max_column
    #print("Trabajador: ", name)
    #print("Semana: ", sheet_obj)

    for k in range(0, len(week)):
        #print("Dia: ", sheet_obj.cell(row=week[k][0], column=week[k][1]).value)
        fecha = sheet_obj.cell(row=week[k][0], column=week[k][1]).value
        for i in range(5, 33):

            cell_obj = sheet_obj.cell(row=week[k][0] + 2, column=i)
            cell_hour = sheet_obj.cell(row=week[k][0], column=i)
            # print(cell_hour.value, end=" ")
            if (cell_obj.value is not None):
                #print(cell_hour.value, cell_obj.value, end=" \n")
                #print(fecha)
                if (cell_obj.value == "F"):
                    new = {

                        'id': len(horario) + 1,
                        'fecha': sheet_obj.cell(row=week[k][0], column=week[k][1]).value,
                        'hora': 'Libre'

                    }

                else:

                    new = {
                        'id': len(horario) + 1,
                        'fecha': sheet_obj.cell(row=week[k][0], column=week[k][1]).value,
                        'hora': cell_hour.value
                    }
                horario.append(new)

        #print("Fin de dia --------------------------")


@app.route('/', methods=['GET'])
def home():
    return jsonify({"message": "Welcome to the Flask backend!"}), 200

@app.route('/file', methods=['GET'])
def archivo():
    global horario
    horario = []
    horas()
    return horario, 200

@socketio.on('message')
def handle_message(msg):
    print('Mensaje recibido:', msg)
    send('Hola desde el servidor Python!', broadcast=True)

# Agregar las rutas a la API
api.add_resource(ItemList, '/items')
api.add_resource(Item, '/items/<int:item_id>', )



if __name__ == '__main__':
    app.run(host='0.0.0.0', port=3000)
